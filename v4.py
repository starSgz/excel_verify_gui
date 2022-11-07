import pymysql
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font
import openpyxl as vb
import os

import warnings
warnings.filterwarnings('ignore')

##单进程使用，添加 网元描述、网元集团的校验 读取校验

class verify_data:
    def __init__(self,excel_url='/',progressbarOne=None,master=None):


        self.db = pymysql.connect(
                host = "localhost", #主机ip
                user = "root", #数据库用户
                password = "123456", #用户对应的密码
                database = "t_wlb_cbwl_km_conf", #对应的数据库
                port = 3306, #数据库端口，默认3306
                charset = 'utf8' #数据库编码
                )

        # 创建游标:游标用于传递python给mysql的命令和mysql返回的内容
        self.cursor = self.db.cursor()
        self.excel_url=excel_url

        self.workbook_b = vb.load_workbook(self.excel_url)
        self.sheet_b = self.workbook_b['Sheet1']
        self.max_rowb = self.sheet_b.max_row

        self.progressbarOne = progressbarOne
        self.master = master
        self.progressbarOne['maximum'] = 1
        self.progressbarOne['value'] = 0

    def excute_getOne(self, sql):
        self.cursor.execute(sql)
        rows = self.cursor.fetchall()
        return rows

    # 从mysql读取数据  只读取一条
    def get_one_data(self,code):
        cmd = '''SELECT
         b.ywzxh_bh,  -- 项目编号
          b.ywzxm_mc,  -- 项目名称 
         b.tr_dx,   -- 投入对象
         b.cb_lx,   -- 成本类型
          b.sy_lb,   -- 事由类别
         c.jjsx,    -- 经济事项
         b.yt,     -- 用途
         c.jjsx_id   -- 经济事项id
        FROM
         ((
         SELECT
          a.ywzxh_bh,
          a.ywzxm_mc,
          a.sy_lb,
          a.tr_dx,
          ( CASE a.sijys_zb_bm WHEN 'CW0961' THEN '低耗品' ELSE '修理费' END ) cb_lx,
          a.tr_dx as yt
         FROM
          `t_wlb_total_list_new` a 
         WHERE
          a.ywzxh_bh = '{}'   -- 变量1：项目id
         ) b
         LEFT JOIN t_wlb_jjsx_conf c ON b.cb_lx = c.cb_type 
         AND b.sy_lb = c.sy_type
				 )
        '''.format(code)

        # print(cmd)
        exe = self.cursor.execute(cmd)  # 执行命令，返回查询的条数
        result = self.cursor.fetchone()  # 查询结果


        # print(cmd)
        # print(result)
        if result == ():  # 没有数据
            print(code)
            # time.sleep(5)
            return [(code, '', '', '', '', '', '', ''),]
        elif result == None:
            print(code)
            # time.sleep(5)
            return [(code, '', '', '', '', '', '', ''),]
        else:
            return list(result)

    def get_mysql_data(self):
        # excel
        csv_data = pd.read_excel(self.excel_url)[['WBS编码', '投入对象', '事由类别', '凭证行项文本摘要', '投入对象', '经济事项']]

        # mysql
        list_mysql = []
        for i in csv_data['WBS编码']:
            print('sssss')
            self.progressbarOne['value']+=1/((self.max_rowb+1-2)*2)
            self.master.update()
            list_mysql.append(self.get_one_data(i))
        mysql_data = pd.DataFrame(list_mysql,columns=['ywzxh_bh', 'ywzxm_mc', 'tr_dx', 'cb_lx', 'sy_lb', 'jjsx', 'yt', 'jjsx_id'])
        if os.path.exists('mysqlData.xlsx'):
            os.remove('mysqlData.xlsx')
            mysql_data.to_excel('mysqlData.xlsx')
        else:
            mysql_data.to_excel('mysqlData.xlsx')


    def verify_excel_data(self):

        self.get_mysql_data()

        workbook_a = vb.load_workbook(r'mysqlData.xlsx') #mysql
        # workbook_b = vb.load_workbook(self.excel_url)             #excel
        # 读取表 总表长度 5812
        sheet_a = workbook_a['Sheet1']  # mysql
        # sheet_b = workbook_b['Sheet1']  # excel

        #获取excel长度
        # max_rowb = sheet_b.max_row


        for i in range(2, self.max_rowb+1):
            self.progressbarOne['value']+=1/((self.max_rowb+1-2)*2)
            self.master.update()
            # WBS编码
            a = sheet_a.cell(i, 2)
            b = self.sheet_b.cell(i, 4)
            if a.value != b.value:
                print(1, i)
                a.fill = PatternFill("solid", fgColor="FF0000")
                a.font = Font(color=colors.BLACK, bold=True)

                b.fill = PatternFill("solid", fgColor="FF0000")
                b.font = Font(color=colors.BLACK, bold=True)

            # 投入对象
            a = sheet_a.cell(i, 4)
            b = self.sheet_b.cell(i, 13)
            if a.value != b.value:
                print(2, i)
                a.fill = PatternFill("solid", fgColor="FF0000")
                a.font = Font(color=colors.BLACK, bold=True)

                b.fill = PatternFill("solid", fgColor="FF0000")
                b.font = Font(color=colors.BLACK, bold=True)


            # 添加所有判断判断集团网元
            cmd_second = '''
            select dywyms,dywybm from t_wlb_xlf_wy_conf c where c.trdx ='{}'
            and dywybm = '{}'
            '''.format(b.value, self.sheet_b.cell(i, 27).value)
            if self.excute_getOne(cmd_second)==():
                self.sheet_b.cell(i, 27).fill = PatternFill("solid", fgColor="FF0000")
                self.sheet_b.cell(i, 27).font = Font(color=colors.BLACK, bold=True)
                self.sheet_b.cell(i, 28).fill = PatternFill("solid", fgColor="FF0000")
                self.sheet_b.cell(i, 28).font = Font(color=colors.BLACK, bold=True)

            # # 添加所有判断判断网元描述
            # cmd_third = '''
            #         select dywyms,dywybm from t_wlb_xlf_wy_conf c where c.trdx ='{}'
            #         and dywybm = '{}'
            #         '''.format(b.value, sheet_b.cell(i, 28).value)
            # print(cmd_third)
            # print(self.excute_getOne(cmd_third))
            #
            # if self.excute_getOne(cmd_third)==():
            #     sheet_b.cell(i, 28).fill = PatternFill("solid", fgColor="FF0000")
            #     sheet_b.cell(i, 28).font = Font(color=colors.BLACK, bold=True)

            # 事由类别
            a = sheet_a.cell(i, 6)
            b = self.sheet_b.cell(i, 12)
            if a.value != b.value:
                print(3, i)
                a.fill = PatternFill("solid", fgColor="FF0000")
                a.font = Font(color=colors.BLACK, bold=True)

                b.fill = PatternFill("solid", fgColor="FF0000")
                b.font = Font(color=colors.BLACK, bold=True)

                ##凭证行项文本摘要  有差异暂时不做校验
                # a=sheet_a.cell(i, 6)
                # b=sheet_b.cell(i, 12)
                # if a.value!=b.value:
                #     print(i)
                #     a.fill = PatternFill("solid", fgColor="#DC143C")
                #     a.font = Font(color=colors.BLACK,bold=True)
                #
                #     b.fill = PatternFill("solid", fgColor="#DC143C")
                #     b.font = Font(color=colors.BLACK,bold=True)

            # 用途
            a = sheet_a.cell(i, 8)
            b = self.sheet_b.cell(i, 13)
            if a.value != b.value:
                print(6, i)
                a.fill = PatternFill("solid", fgColor="FF0000")
                a.font = Font(color=colors.BLACK, bold=True)

                b.fill = PatternFill("solid", fgColor="FF0000")
                b.font = Font(color=colors.BLACK, bold=True)

            # 经济事项
            a = sheet_a.cell(i, 9)
            b = self.sheet_b.cell(i, 16)
            if a.value != b.value:
                print(7, i)
                a.fill = PatternFill("solid", fgColor="FF0000")
                a.font = Font(color=colors.BLACK, bold=True)

                b.fill = PatternFill("solid", fgColor="FF0000")
                b.font = Font(color=colors.BLACK, bold=True)

        # workbook_a.save(r'mysqlData.xlsx')
        self.workbook_b.save(self.excel_url)
        #校验完成后删除mysqlData文件
        # os.remove('mysqlData.xlsx')
        self.cursor.close()  # 关闭游标
        self.db.close()  # 关闭链接


# if __name__ == '__main__':
#     test = verify_data()
#     # test.get_one_data('220AGDGZ0001-001')
#     test.excel_url = '修理费相关科目列账明细-原始导出.xlsx'
#     # test.get_mysql_data()
#     # print(test.pandas_mysql)
#
#     test.verify_excel_data()


