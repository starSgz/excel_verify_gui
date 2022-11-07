import pymysql
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font
import openpyxl as vb
from dbutils.pooled_db import PooledDB
from concurrent.futures import ThreadPoolExecutor, as_completed
import datetime
import threading
import warnings
warnings.filterwarnings('ignore')
####################################v2 pandas 废弃版本
class verify_data:
    def __init__(self):
        #数据库连接池
        self.__pool = PooledDB(creator=pymysql,
                               mincached=1,
                               maxcached=5,
                               maxshared=10,
                               maxconnections=5,
                               maxusage=5,
                               blocking=True,
                               user="root",
                               passwd="123456",
                               db="t_wlb_cbwl_km_conf",
                               host="localhost",
                               port=3306,
                               charset='utf8',
                               )

        # 创建游标:游标用于传递python给mysql的命令和mysql返回的内容
        self.excel_url="/"
        self.pandas_mysql=pd.DataFrame(columns = ['ywzxh_bh', 'ywzxm_mc', 'tr_dx', 'cb_lx', 'sy_lb', 'jjsx', 'yt', 'jjsx_id','dywyms','dywybm'],index=None)
        self.lock =threading.RLock()
    # 计算时间的装饰器
    def task_content_time(func):
        start_time = datetime.datetime.now()
        def wrapper(*args, **kwargs):
            func(*args, **kwargs)
            end_time = datetime.datetime.now()
            take_time = end_time - start_time
            print(func.__name__,"任务总共花费时间:", take_time)
        return wrapper

    #获取数据库连接池连接
    def get_conn(self):
        conn = self.__pool.connection()  # 从连接池获取一个链接
        cursor = conn.cursor()
        return conn, cursor

    #关闭连接
    @staticmethod
    def dispose(cursor, conn):
        cursor.close()
        conn.close()

    #执行一条sql
    def excute_getOne(self, sql):
        conn, cursor = self.get_conn()
        cursor.execute(sql)
        rows = cursor.fetchall()

        self.dispose(cursor, conn)
        return rows

    # 从mysql读取数据  只读取一条
    def get_one_data(self,code):
        cmd = '''
        SELECT
         b.ywzxh_bh,  -- 项目编号
          b.ywzxm_mc,  -- 项目名称 
         b.tr_dx,   -- 投入对象
         b.cb_lx,   -- 成本类型
          b.sy_lb,   -- 事由类别
         c.jjsx,    -- 经济事项
         b.yt,     -- 用途
         c.jjsx_id,   -- 经济事项id
				 d.dywyms,
				 d.dywybm
        FROM
         ((
         SELECT
          a.ywzxh_bh,
          a.ywzxm_mc,
          a.sy_lb,
          a.tr_dx,
          ( CASE a.sijys_zb_bm WHEN 'CW0961' THEN '低耗品' ELSE '修理费' END ) cb_lx,
          a.tr_dx as yt
         FROM
          `t_wlb_total_list_new` a 
         WHERE
          a.ywzxh_bh = '{}'   -- 变量1：项目id
         ) b
         LEFT JOIN t_wlb_jjsx_conf c ON b.cb_lx = c.cb_type 
         AND b.sy_lb = c.sy_type
				 )
				 LEFT JOIN t_wlb_xlf_wy_conf d ON b.tr_dx = d.trdx
				 '''.format(code)

        result = self.excute_getOne(cmd)

        # print(cmd)
        # print(result)

        if result == ():  # 没有数据
            print(cmd)
            return [(code, '', '', '', '', '', '', '','',''),]

        elif result == None:
            print(cmd)
            return [(code, '', '', '', '', '', '', '','',''),]

        else:
            return list(result)

    @task_content_time
    def get_mysql_data(self):
        '''
        多线程锁 不如直接单线程跑
        '''

        #线程池
        pool = ThreadPoolExecutor(max_workers=10)#执行mysql的线程池
        pool1 = ThreadPoolExecutor(max_workers=10)#执行mysql的线程池

        # excel
        csv_data = pd.read_excel(self.excel_url)
        task_list = []
        task_list1 = []

        # mysql
        for i in csv_data['WBS编码']:
            # print(i)
            future = pool.submit(self.get_one_data,i)
            task_list.append(future)

        for future in as_completed(task_list):
            data  = future.result()
            task_next = pool1.submit(self.append_pandas,data)
            task_list1.append(task_next)

        for i in as_completed(task_list1):
            data = i.result()



        # self.pandas_mysql.to_excel('mysqlDataMany.xlsx')

    def append_pandas(self,data):
        for data_one in data:
            data_one = list(data_one)
            self.lock.acquire()
            length = len(self.pandas_mysql)
            self.pandas_mysql.loc[length] = data_one
            self.lock.release()
            print(len(self.pandas_mysql))
        # print(self.pandas_mysql)

    def verify_excel_data(self):

        self.get_mysql_data()

        excel_data = vb.load_workbook(self.excel_url)
        excel_sheet = excel_data['Sheet1']  # excel

        max_row = excel_sheet.max_row

        for i in range (2,max_row+1):
            #WBS编码
            excel_verify_WBS = excel_sheet.cell(i, 4)
            mysql_verify = self.pandas_mysql[(self.pandas_mysql['ywzxh_bh']==excel_verify_WBS.value)]
            if excel_verify_WBS.value != mysql_verify['ywzxh_bh'].iloc[0]:
                excel_verify_WBS.fill = PatternFill("solid", fgColor="FF0000")
                excel_verify_WBS.font = Font(color=colors.BLACK, bold=True)

            #投入对象
            excel_verify_trdx = excel_sheet.cell(i, 13)
            if excel_verify_trdx.value!= mysql_verify['tr_dx'].iloc[0]:
                excel_verify_trdx.fill =PatternFill("solid", fgColor="FF0000")
                excel_verify_trdx.font =Font(color=colors.BLACK, bold=True)

            #事由类别
            excel_verify_sylb = excel_sheet.cell(i, 12)
            if excel_verify_sylb.value != mysql_verify['sy_lb'].iloc[0]:
                excel_verify_sylb.fill =PatternFill("solid", fgColor="FF0000")
                excel_verify_sylb.font =Font(color=colors.BLACK, bold=True)

            # 用途
            excel_verify_yt = excel_sheet.cell(i, 13)
            if excel_verify_yt.value != mysql_verify['yt'].iloc[0]:
                excel_verify_yt.fill = PatternFill("solid", fgColor="FF0000")
                excel_verify_yt.font = Font(color=colors.BLACK, bold=True)


            # 经济事项
            excel_verify_jjsx = excel_sheet.cell(i, 16)
            if excel_verify_jjsx.value != mysql_verify['jjsx'].iloc[0]:
                excel_verify_jjsx.fill = PatternFill("solid", fgColor="FF0000")
                excel_verify_jjsx.font = Font(color=colors.BLACK, bold=True)

            # 集团网元
            excel_verify_jtwy = excel_sheet.cell(i, 27)
            # print(excel_verify_jtwy.value)
            # print(any(mysql_verify['dywybm']==excel_verify_jtwy.value))
            if any(mysql_verify['dywybm']==excel_verify_jtwy.value):
                pass
            else:
                # print(1)
                excel_verify_jtwy.fill = PatternFill("solid", fgColor="FF0000")
                excel_verify_jtwy.font = Font(color=colors.BLACK, bold=True)

            #网元描述
            excel_verify_wyms = excel_sheet.cell(i, 28)
            if any(mysql_verify['dywyms']==excel_verify_wyms.value):
                pass
            else:
                excel_verify_wyms.fill = PatternFill("solid", fgColor="FF0000")
                excel_verify_wyms.font = Font(color=colors.BLACK, bold=True)
            print(excel_verify_WBS)
            excel_data.save(self.excel_url)





if __name__ == '__main__':
    test = verify_data()
    # test.get_one_data('220AGDGZ0001-001')
    test.excel_url='修理费相关科目列账明细-原始导出.xlsx'
    test.get_mysql_data()
    # print(test.pandas_mysql)

    # test.verify_excel_data()


